"""
Export Service (v4 — fixed Kurdish/Arabic text in PDF)
Registers an Arabic-supporting font so Kurdish text renders properly.
"""

import os
import re
from datetime import datetime
from typing import List, Dict, Optional

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


EXPORT_DIR = "exports"
os.makedirs(EXPORT_DIR, exist_ok=True)


# ===== BIDI/RESHAPING SUPPORT =====
try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    HAS_BIDI = True
except ImportError:
    HAS_BIDI = False


# ===== FONT REGISTRATION =====

def _register_arabic_font() -> Optional[str]:
    candidates = [
        ("Tahoma", [
            "C:/Windows/Fonts/tahoma.ttf",
            "C:\\Windows\\Fonts\\tahoma.ttf",
        ]),
        ("Arial", [
            "C:/Windows/Fonts/arial.ttf",
            "C:\\Windows\\Fonts\\arial.ttf",
        ]),
        ("SegoeUI", [
            "C:/Windows/Fonts/segoeui.ttf",
            "C:\\Windows\\Fonts\\segoeui.ttf",
        ]),
        ("FreeSerif", [
            "/usr/share/fonts/truetype/freefont/FreeSerif.ttf",
        ]),
        ("DejaVuSans", [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]),
    ]

    for font_name, paths in candidates:
        for path in paths:
            try:
                if os.path.exists(path):
                    pdfmetrics.registerFont(TTFont(font_name, path))
                    bold_path = path.replace(".ttf", "bd.ttf")
                    if not os.path.exists(bold_path):
                        bold_path = path.replace(".ttf", "-Bold.ttf")
                    if os.path.exists(bold_path):
                        try:
                            pdfmetrics.registerFont(TTFont(font_name + "Bd", bold_path))
                        except:
                            pass
                    return font_name
            except:
                continue
    return None


_ARABIC_FONT = _register_arabic_font()
_FN = _ARABIC_FONT or "Helvetica"
_FNB = (_ARABIC_FONT + "Bd") if _ARABIC_FONT else "Helvetica-Bold"

try:
    pdfmetrics.getFont(_FNB)
except:
    _FNB = _FN


def _has_arabic(text):
    if not text:
        return False
    return bool(re.search(r'[\u0600-\u06FF\u0750-\u077F\uFB50-\uFDFF\uFE70-\uFEFF]', text))


def _reshape(text):
    if not text or not _has_arabic(text):
        return text or ""
    if HAS_BIDI:
        try:
            return get_display(arabic_reshaper.reshape(text))
        except:
            return text
    return text


def _safe(text):
    if not text:
        return "-"
    if _has_arabic(text):
        if _ARABIC_FONT:
            return _reshape(text)
        latin = re.sub(r'[\u0600-\u06FF\u0750-\u077F\uFB50-\uFDFF\uFE70-\uFEFF]+', '', text).strip()
        return latin if latin else "(see web interface)"
    return text


def export_to_pdf(job_title: str, ranked_cvs: List[Dict], summary: Dict) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXPORT_DIR, f"cv_results_{timestamp}.pdf")

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()

    s_title = ParagraphStyle('T', parent=styles['Title'], fontName=_FNB, fontSize=18, spaceAfter=6)
    s_sub = ParagraphStyle('S', parent=styles['Normal'], fontName=_FN, fontSize=10, textColor=colors.grey, spaceAfter=16)
    s_h2 = ParagraphStyle('H', parent=styles['Heading2'], fontName=_FNB, fontSize=13, spaceAfter=8, spaceBefore=14)
    s_body = ParagraphStyle('B', parent=styles['Normal'], fontName=_FN, fontSize=9, spaceAfter=4)
    s_bold = ParagraphStyle('BB', parent=styles['Normal'], fontName=_FNB, fontSize=9, spaceAfter=4)
    s_foot = ParagraphStyle('F', parent=styles['Normal'], fontName=_FN, fontSize=7, textColor=colors.grey)

    el = []

    el.append(Paragraph("CV Processing Results Report", s_title))
    el.append(Paragraph(f"Job: {_safe(job_title)} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", s_sub))

    # Summary table
    el.append(Paragraph("Summary", s_h2))
    st = [
        ["Total CVs", "Processed", "Best Match", "Average Score"],
        [str(summary.get("total_cvs", 0)), str(summary.get("completed", 0)),
         f"{summary.get('highest_score', 0)*100:.1f}%", f"{summary.get('average_score', 0)*100:.1f}%"],
    ]
    t = Table(st, colWidths=[120, 100, 100, 100])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1b6b5a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), _FNB),
        ('FONTNAME', (0,1), (-1,-1), _FN),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f0f9f6')),
    ]))
    el.append(t)
    el.append(Spacer(1, 10))

    # Ranked results
    el.append(Paragraph("Ranked Results", s_h2))
    td = [["Rank", "Filename", "Language", "Score", "Label"]]
    for cv in ranked_cvs:
        td.append([
            str(cv.get("rank", "")),
            cv.get("filename", "")[:35],
            cv.get("detected_language", ""),
            f"{cv.get('similarity_percentage', 0):.1f}%",
            cv.get("similarity_label", ""),
        ])

    rt = Table(td, colWidths=[35, 180, 65, 60, 95])
    ts = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1b6b5a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), _FNB),
        ('FONTNAME', (0,1), (-1,-1), _FN),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (3,0), (3,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]
    for i, cv in enumerate(ranked_cvs, 1):
        sc = cv.get("similarity_percentage", 0)
        if sc >= 70: ts.append(('TEXTCOLOR', (3,i), (3,i), colors.HexColor('#16a34a')))
        elif sc >= 50: ts.append(('TEXTCOLOR', (3,i), (3,i), colors.HexColor('#1b6b5a')))
        elif sc >= 30: ts.append(('TEXTCOLOR', (3,i), (3,i), colors.HexColor('#d97706')))
        else: ts.append(('TEXTCOLOR', (3,i), (3,i), colors.HexColor('#dc2626')))
    rt.setStyle(TableStyle(ts))
    el.append(rt)

    # Top candidate details
    el.append(Spacer(1, 16))
    el.append(Paragraph("Top Candidate Details", s_h2))

    for cv in ranked_cvs[:5]:
        el.append(Paragraph(f"<b>#{cv.get('rank')} - {cv.get('filename', '')}</b>", s_bold))

        lines = []
        if cv.get("parsed_name"):
            lines.append(f"Name: {_safe(cv['parsed_name'])}")
        if cv.get("parsed_email"):
            lines.append(f"Email: {cv['parsed_email']}")
        if cv.get("parsed_degree"):
            lines.append(f"Degree: {_safe(cv['parsed_degree'])}")
        if cv.get("parsed_skills"):
            lines.append(f"Skills: {', '.join(cv['parsed_skills'][:10])}")
        if cv.get("matched_keywords"):
            kws = [_safe(k.get("keyword", "")) for k in cv["matched_keywords"][:8]]
            lines.append(f"Matched Keywords: {', '.join(kws)}")

        lines.append(
            f"Scores: CrossLang {cv.get('crosslang_score',0)*100:.1f}% | "
            f"TF-IDF {cv.get('tfidf_score',0)*100:.1f}% | "
            f"Combined {cv.get('similarity_percentage',0):.1f}%"
        )
        for line in lines:
            el.append(Paragraph(f"  {line}", s_body))
        el.append(Spacer(1, 6))

    el.append(Spacer(1, 20))
    el.append(Paragraph("Generated by Automatic CV Processing System | BSc Graduation Project", s_foot))

    doc.build(el)
    return filepath


def export_to_excel(job_title: str, ranked_cvs: List[Dict], summary: Dict) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXPORT_DIR, f"cv_results_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='1B6B5A')
    df = Font(name='Arial', size=10)
    bd = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

    ws['A1'] = 'CV Processing Results'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1B6B5A')
    ws['A2'] = f'Job: {job_title}'
    ws['A2'].font = Font(name='Arial', size=10, color='666666')
    ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A3'].font = Font(name='Arial', size=10, color='666666')

    for c, h in enumerate(['Metric', 'Value'], 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')

    rows = [
        ('Total CVs', summary.get('total_cvs', 0)),
        ('Processed', summary.get('completed', 0)),
        ('Errors', summary.get('errors', 0)),
        ('Best Match', f"{summary.get('highest_score', 0)*100:.1f}%"),
        ('Lowest Match', f"{summary.get('lowest_score', 0)*100:.1f}%"),
        ('Average Score', f"{summary.get('average_score', 0)*100:.1f}%"),
    ]
    for i, (m, v) in enumerate(rows, 6):
        ws.cell(row=i, column=1, value=m).font = df
        ws.cell(row=i, column=2, value=v).font = df
        ws.cell(row=i, column=1).border = bd
        ws.cell(row=i, column=2).border = bd

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    # Sheet 2
    ws2 = wb.create_sheet("Ranked Results")
    headers = ['Rank', 'Filename', 'Name', 'Email', 'Degree', 'Language',
               'TF-IDF %', 'CrossLang %', 'Combined %', 'Label', 'Matched Keywords', 'Skills']

    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = bd

    for i, cv in enumerate(ranked_cvs, 2):
        sk = cv.get('parsed_skills', [])
        mk = cv.get('matched_keywords', [])
        vals = [
            cv.get('rank',''), cv.get('filename',''), cv.get('parsed_name',''),
            cv.get('parsed_email',''), cv.get('parsed_degree',''), cv.get('detected_language',''),
            round(cv.get('tfidf_score',0)*100, 1), round(cv.get('crosslang_score',0)*100, 1),
            round(cv.get('similarity_percentage',0), 1), cv.get('similarity_label',''),
            ', '.join([k.get('keyword','') for k in mk[:10]]),
            ', '.join(sk[:15]),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font = df; cell.border = bd; cell.alignment = Alignment(vertical='center')

        sc_cell = ws2.cell(row=i, column=9)
        sc = cv.get('similarity_percentage', 0)
        if sc >= 70: sc_cell.font = Font(name='Arial', size=10, bold=True, color='16A34A')
        elif sc >= 50: sc_cell.font = Font(name='Arial', size=10, bold=True, color='1B6B5A')
        elif sc >= 30: sc_cell.font = Font(name='Arial', size=10, color='D97706')
        else: sc_cell.font = Font(name='Arial', size=10, color='DC2626')

        if i % 2 == 0:
            for c in range(1, len(headers)+1):
                ws2.cell(row=i, column=c).fill = PatternFill('solid', fgColor='F8F9FA')

    cols = ['A','B','C','D','E','F','G','H','I','J','K','L']
    widths = [6, 28, 20, 24, 14, 10, 10, 10, 10, 14, 30, 30]
    for letter, w in zip(cols, widths):
        ws2.column_dimensions[letter].width = w

    ws2.auto_filter.ref = f"A1:L{len(ranked_cvs)+1}"
    wb.save(filepath)
    return filepath
